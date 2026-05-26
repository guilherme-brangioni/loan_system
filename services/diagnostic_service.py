import os
import tempfile
from typing import Any

from flask import current_app
from openpyxl import load_workbook

from services.app_setting_service import AppSettingService
from services.backup_service import BackupService


class DiagnosticService:
    """
    Serviço de diagnóstico do sistema.

    Apenas verifica caminhos, arquivos e configurações.
    Não altera dados do banco nem das planilhas.
    """

    STATUS_OK = "OK"
    STATUS_WARNING = "AVISO"
    STATUS_ERROR = "ERRO"

    @staticmethod
    def run_all_checks() -> list[dict[str, Any]]:
        """
        Executa todos os diagnósticos.
        """

        checks = []

        checks.append(DiagnosticService.check_database())
        checks.append(DiagnosticService.check_app_base_url())
        checks.append(DiagnosticService.check_generated_pdf_dir())
        checks.append(DiagnosticService.check_equipment_image_dir())
        checks.append(DiagnosticService.check_backup_dir())
        checks.append(DiagnosticService.check_automatic_backup())
        checks.append(DiagnosticService.check_logo_path())
        checks.append(DiagnosticService.check_inventory_excel_file())
        checks.append(DiagnosticService.check_inventory_excel_sheet())
        checks.append(DiagnosticService.check_movements_excel_path())

        return checks

    @staticmethod
    def _make_result(
        name: str,
        status: str,
        message: str,
        value: str = "",
        details: str = "",
    ) -> dict[str, Any]:
        return {
            "name": name,
            "status": status,
            "message": message,
            "value": value,
            "details": details,
        }

    @staticmethod
    def _can_write_to_dir(directory: str) -> bool:
        """
        Testa escrita criando e removendo um arquivo temporário.
        """

        if not directory:
            return False

        if not os.path.exists(directory):
            return False

        if not os.path.isdir(directory):
            return False

        try:
            fd, temp_path = tempfile.mkstemp(
                prefix="diagnostic_",
                suffix=".tmp",
                dir=directory,
            )

            os.close(fd)
            os.remove(temp_path)

            return True

        except Exception:
            return False

    @staticmethod
    def _ensure_dir_status(directory: str, name: str) -> dict[str, Any]:
        """
        Diagnóstico genérico de pasta.
        """

        if not directory:
            return DiagnosticService._make_result(
                name=name,
                status=DiagnosticService.STATUS_ERROR,
                message="Caminho não configurado.",
                value="",
            )

        if not os.path.exists(directory):
            return DiagnosticService._make_result(
                name=name,
                status=DiagnosticService.STATUS_WARNING,
                message="Pasta não existe. O sistema pode tentar criá-la quando necessário.",
                value=directory,
            )

        if not os.path.isdir(directory):
            return DiagnosticService._make_result(
                name=name,
                status=DiagnosticService.STATUS_ERROR,
                message="O caminho existe, mas não é uma pasta.",
                value=directory,
            )

        if not DiagnosticService._can_write_to_dir(directory):
            return DiagnosticService._make_result(
                name=name,
                status=DiagnosticService.STATUS_ERROR,
                message="Pasta encontrada, mas sem permissão de escrita.",
                value=directory,
            )

        return DiagnosticService._make_result(
            name=name,
            status=DiagnosticService.STATUS_OK,
            message="Pasta encontrada e com permissão de escrita.",
            value=directory,
        )

    @staticmethod
    def check_database() -> dict[str, Any]:
        """
        Verifica o banco SQLite.
        """

        try:
            database_path = BackupService.get_database_path()

            if not os.path.exists(database_path):
                return DiagnosticService._make_result(
                    name="Banco de dados",
                    status=DiagnosticService.STATUS_ERROR,
                    message="Arquivo do banco não encontrado.",
                    value=database_path,
                )

            size_mb = os.path.getsize(database_path) / 1024 / 1024

            return DiagnosticService._make_result(
                name="Banco de dados",
                status=DiagnosticService.STATUS_OK,
                message=f"Banco encontrado. Tamanho: {size_mb:.2f} MB.",
                value=database_path,
            )

        except Exception as exc:
            return DiagnosticService._make_result(
                name="Banco de dados",
                status=DiagnosticService.STATUS_ERROR,
                message="Erro ao verificar banco de dados.",
                details=str(exc),
            )

    @staticmethod
    def check_app_base_url() -> dict[str, Any]:
        """
        Verifica APP_BASE_URL usado em QR Code e links.
        """

        base_url = AppSettingService.get(
            "APP_BASE_URL",
            current_app.config.get("APP_BASE_URL", ""),
        )

        base_url = str(base_url or "").strip()

        if not base_url:
            return DiagnosticService._make_result(
                name="URL base do sistema",
                status=DiagnosticService.STATUS_ERROR,
                message="APP_BASE_URL não configurado.",
                value="",
            )

        if "127.0.0.1" in base_url or "localhost" in base_url.lower():
            return DiagnosticService._make_result(
                name="URL base do sistema",
                status=DiagnosticService.STATUS_WARNING,
                message="URL local. QR Code funcionará apenas no próprio computador.",
                value=base_url,
            )

        return DiagnosticService._make_result(
            name="URL base do sistema",
            status=DiagnosticService.STATUS_OK,
            message="URL base configurada.",
            value=base_url,
        )

    @staticmethod
    def check_generated_pdf_dir() -> dict[str, Any]:
        directory = current_app.config.get("GENERATED_PDF_DIR", "")

        return DiagnosticService._ensure_dir_status(
            directory=directory,
            name="Pasta de PDFs gerados",
        )

    @staticmethod
    def check_equipment_image_dir() -> dict[str, Any]:
        directory = current_app.config.get("EQUIPMENT_IMAGE_UPLOAD_DIR", "")

        return DiagnosticService._ensure_dir_status(
            directory=directory,
            name="Pasta de imagens dos equipamentos",
        )

    @staticmethod
    def check_backup_dir() -> dict[str, Any]:
        backup_dir = AppSettingService.get(
            "BACKUP_DIR",
            current_app.config.get("BACKUP_DIR", ""),
        )

        backup_dir = str(backup_dir or "").strip()

        return DiagnosticService._ensure_dir_status(
            directory=backup_dir,
            name="Pasta de backups",
        )

    @staticmethod
    def check_automatic_backup() -> dict[str, Any]:
        """
        Verifica se existe backup automático.
        """

        try:
            backup = BackupService.get_automatic_backup()

            if not backup:
                return DiagnosticService._make_result(
                    name="Backup automático",
                    status=DiagnosticService.STATUS_WARNING,
                    message="Nenhum backup automático encontrado ainda.",
                    value="",
                )

            return DiagnosticService._make_result(
                name="Backup automático",
                status=DiagnosticService.STATUS_OK,
                message="Backup automático encontrado.",
                value=backup["path"],
                details=(
                    "Última atualização: "
                    + backup["created_at"].strftime("%d/%m/%Y %H:%M:%S")
                ),
            )

        except Exception as exc:
            return DiagnosticService._make_result(
                name="Backup automático",
                status=DiagnosticService.STATUS_ERROR,
                message="Erro ao verificar backup automático.",
                details=str(exc),
            )

    @staticmethod
    def check_logo_path() -> dict[str, Any]:
        logo_path = AppSettingService.get(
            "LOGO_PATH",
            current_app.config.get("LOGO_PATH", ""),
        )

        logo_path = str(logo_path or "").strip()

        if not logo_path:
            return DiagnosticService._make_result(
                name="Logo do PDF",
                status=DiagnosticService.STATUS_WARNING,
                message="Logo não configurada. O PDF será gerado sem logo.",
                value="",
            )

        if not os.path.exists(logo_path):
            return DiagnosticService._make_result(
                name="Logo do PDF",
                status=DiagnosticService.STATUS_WARNING,
                message="Arquivo de logo não encontrado. O PDF será gerado sem logo.",
                value=logo_path,
            )

        return DiagnosticService._make_result(
            name="Logo do PDF",
            status=DiagnosticService.STATUS_OK,
            message="Arquivo de logo encontrado.",
            value=logo_path,
        )

    @staticmethod
    def _get_inventory_file_path() -> str:
        return str(
            AppSettingService.get(
                "EXCEL_INVENTORY_FILE",
                current_app.config.get("EXCEL_INVENTORY_FILE", ""),
            )
            or ""
        ).strip()

    @staticmethod
    def _get_inventory_sheet_name() -> str:
        return str(
            AppSettingService.get(
                "EXCEL_INVENTORY_SHEET",
                current_app.config.get("EXCEL_INVENTORY_SHEET", ""),
            )
            or ""
        ).strip()

    @staticmethod
    def check_inventory_excel_file() -> dict[str, Any]:
        file_path = DiagnosticService._get_inventory_file_path()

        if not file_path:
            return DiagnosticService._make_result(
                name="Planilha de inventário",
                status=DiagnosticService.STATUS_ERROR,
                message="Caminho da planilha de inventário não configurado.",
                value="",
            )

        if not os.path.exists(file_path):
            return DiagnosticService._make_result(
                name="Planilha de inventário",
                status=DiagnosticService.STATUS_ERROR,
                message="Arquivo da planilha de inventário não encontrado.",
                value=file_path,
            )

        return DiagnosticService._make_result(
            name="Planilha de inventário",
            status=DiagnosticService.STATUS_OK,
            message="Arquivo da planilha de inventário encontrado.",
            value=file_path,
        )

    @staticmethod
    def check_inventory_excel_sheet() -> dict[str, Any]:
        file_path = DiagnosticService._get_inventory_file_path()
        sheet_name = DiagnosticService._get_inventory_sheet_name()

        if not file_path or not os.path.exists(file_path):
            return DiagnosticService._make_result(
                name="Aba da planilha de inventário",
                status=DiagnosticService.STATUS_WARNING,
                message="Não foi possível verificar a aba porque o arquivo não foi encontrado.",
                value=sheet_name,
            )

        if not sheet_name:
            return DiagnosticService._make_result(
                name="Aba da planilha de inventário",
                status=DiagnosticService.STATUS_ERROR,
                message="Nome da aba não configurado.",
                value="",
            )

        workbook = None

        try:
            workbook = load_workbook(
                file_path,
                read_only=True,
                data_only=True,
            )

            if sheet_name not in workbook.sheetnames:
                return DiagnosticService._make_result(
                    name="Aba da planilha de inventário",
                    status=DiagnosticService.STATUS_ERROR,
                    message="Aba não encontrada na planilha.",
                    value=sheet_name,
                    details="Abas disponíveis: " + ", ".join(workbook.sheetnames),
                )

            return DiagnosticService._make_result(
                name="Aba da planilha de inventário",
                status=DiagnosticService.STATUS_OK,
                message="Aba encontrada na planilha.",
                value=sheet_name,
            )

        except Exception as exc:
            return DiagnosticService._make_result(
                name="Aba da planilha de inventário",
                status=DiagnosticService.STATUS_ERROR,
                message="Erro ao abrir planilha para verificar aba.",
                value=file_path,
                details=str(exc),
            )

        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def check_movements_excel_path() -> dict[str, Any]:
        file_path = AppSettingService.get(
            "EXCEL_MOVEMENTS_FILE",
            current_app.config.get("EXCEL_MOVEMENTS_FILE", ""),
        )

        file_path = str(file_path or "").strip()

        if not file_path:
            return DiagnosticService._make_result(
                name="Planilha de movimentações Power BI",
                status=DiagnosticService.STATUS_ERROR,
                message="Caminho da planilha de movimentações não configurado.",
                value="",
            )

        folder = os.path.dirname(file_path)

        if not folder:
            return DiagnosticService._make_result(
                name="Planilha de movimentações Power BI",
                status=DiagnosticService.STATUS_WARNING,
                message="Caminho sem pasta definida. O arquivo será criado na pasta atual.",
                value=file_path,
            )

        folder_result = DiagnosticService._ensure_dir_status(
            directory=folder,
            name="Pasta da planilha de movimentações",
        )

        if folder_result["status"] != DiagnosticService.STATUS_OK:
            return DiagnosticService._make_result(
                name="Planilha de movimentações Power BI",
                status=folder_result["status"],
                message=folder_result["message"],
                value=file_path,
                details=folder_result.get("details", ""),
            )

        if os.path.exists(file_path):
            return DiagnosticService._make_result(
                name="Planilha de movimentações Power BI",
                status=DiagnosticService.STATUS_OK,
                message="Arquivo encontrado e pasta acessível.",
                value=file_path,
            )

        return DiagnosticService._make_result(
            name="Planilha de movimentações Power BI",
            status=DiagnosticService.STATUS_WARNING,
            message="Arquivo ainda não existe, mas a pasta está acessível. O sistema poderá criá-lo.",
            value=file_path,
        )