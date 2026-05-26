import os
from datetime import datetime
from typing import Any

from flask import current_app
from openpyxl import Workbook, load_workbook

from services.audit_service import AuditService
from services.app_setting_service import AppSettingService


class ExcelMovementService:
    """
    Exporta movimentações de empréstimos para Excel.

    Essa planilha pode ser usada pelo Power BI.

    Regra:
    - nunca apaga histórico;
    - cada evento gera uma nova linha;
    - se a exportação falhar, não desfaz a operação no banco.
    """

    COLUMNS = [
        "DATA_REGISTRO",
        "TIPO_MOVIMENTACAO",
        "REALIZADO_POR",
        "OBSERVACAO",

        "LOAN_ID",
        "NUMERO_CONTROLE",
        "STATUS_EMPRESTIMO",

        "SOLICITANTE",
        "MATRICULA",
        "EMAIL_SOLICITANTE",
        "TELEFONE",
        "GERENCIA",
        "REGIONAL_SOLICITANTE",
        "EQUIPE",

        "APROVADOR",
        "EMAIL_APROVADOR",

        "RESPONSAVEL_COLETA_ENTREGA",
        "EMAIL_RESPONSAVEL_COLETA_ENTREGA",

        "DATA_EMPRESTIMO",
        "DATA_PREVISTA_DEVOLUCAO",
        "DATA_REAL_DEVOLUCAO",
        "LOCAL_UTILIZACAO",

        "LOAN_ITEM_ID",
        "STATUS_ITEM",

        "EQUIPMENT_ID",
        "CODIGO_INTERNO",
        "TIPO_EQUIPAMENTO",
        "FABRICANTE",
        "MODELO",
        "PATRIMONIO",
        "CODIGO_EQUIPAMENTO",
        "NUMERO_SERIE",
        "STATUS_EQUIPAMENTO",
        "EQUIPAMENTO_VALIDADO",
        "REGIONAL_EQUIPAMENTO",
        "LOCAL_ARMAZENAGEM",
        "SUBESTACAO_ORIGEM",
    ]

    @staticmethod
    def _format_datetime(value) -> str:
        if not value:
            return ""

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")

        return str(value)

    @staticmethod
    def _get_loan_items(loan: Any) -> list:
        items = getattr(loan, "items", None)

        if not items:
            return []

        return list(items)

    @staticmethod
    def _load_or_create_workbook():
        file_path = ExcelMovementService._get_movements_file_path()

        folder = os.path.dirname(file_path)

        if folder:
            os.makedirs(folder, exist_ok=True)

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        return workbook, file_path

    @staticmethod
    def _get_or_create_sheet(workbook):
        sheet_name = ExcelMovementService._get_movements_sheet_name()

        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

        # Se for uma planilha nova com aba padrão vazia, renomeia.
        active_sheet = workbook.active

        if active_sheet.max_row == 1 and active_sheet["A1"].value is None:
            active_sheet.title = sheet_name
            return active_sheet

        return workbook.create_sheet(sheet_name)

    @staticmethod
    def _ensure_header(sheet) -> None:
        first_row_empty = (
            sheet.max_row == 1
            and all(cell.value is None for cell in sheet[1])
        )

        if first_row_empty:
            for col_index, column_name in enumerate(
                ExcelMovementService.COLUMNS,
                start=1,
            ):
                sheet.cell(row=1, column=col_index).value = column_name

    @staticmethod
    def _build_row(
        loan,
        item,
        movement_type: str,
        performed_by: str,
        notes: str,
    ) -> list:
        user = getattr(loan, "user", None)
        approver = getattr(loan, "approver", None)
        equipment = getattr(item, "equipment", None) if item else None

        return [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            movement_type,
            performed_by or "",
            notes or "",

            getattr(loan, "id", ""),
            getattr(loan, "numero_controle", ""),
            getattr(loan, "status", ""),

            getattr(user, "nome", "") if user else "",
            getattr(user, "matricula", "") if user else "",
            getattr(user, "email", "") if user else "",
            getattr(user, "telefone", "") if user else "",
            getattr(user, "gerencia", "") if user else "",
            getattr(user, "regional", "") if user else "",
            getattr(user, "equipe", "") if user else "",

            getattr(approver, "nome", "") if approver else "",
            getattr(approver, "email", "") if approver else "",

            getattr(loan, "responsavel_entrega_nome", ""),
            getattr(loan, "responsavel_entrega_email", ""),

            ExcelMovementService._format_datetime(
                getattr(loan, "data_emprestimo", None)
            ),
            ExcelMovementService._format_datetime(
                getattr(loan, "data_prevista_devolucao", None)
            ),
            ExcelMovementService._format_datetime(
                getattr(loan, "data_real_devolucao", None)
            ),
            getattr(loan, "local_utilizacao", ""),

            getattr(item, "id", "") if item else "",
            getattr(item, "status", "") if item else "",

            getattr(equipment, "id", "") if equipment else "",
            getattr(equipment, "codigo_interno", "") if equipment else "",
            getattr(equipment, "tipo_equipamento", "") if equipment else "",
            getattr(equipment, "fabricante", "") if equipment else "",
            getattr(equipment, "modelo", "") if equipment else "",
            getattr(equipment, "patrimonio", "") if equipment else "",
            getattr(equipment, "codigo_equipamento", "") if equipment else "",
            getattr(equipment, "serial", "") if equipment else "",
            getattr(equipment, "status", "") if equipment else "",
            "SIM" if equipment and getattr(equipment, "validado", False) else "NAO",
            getattr(equipment, "regional", "") if equipment else "",
            getattr(equipment, "local_armazenagem", "") if equipment else "",
            getattr(equipment, "subestacao_origem", "") if equipment else "",
        ]

    @staticmethod
    def append_loan_movement(
        loan,
        movement_type: str,
        performed_by: str = "",
        notes: str = "",
    ) -> dict:
        """
        Registra movimentação no Excel.

        Retorna:
        {
            "success": True,
            "rows": 2,
            "file": "..."
        }
        """

        workbook = None

        try:
            workbook, file_path = ExcelMovementService._load_or_create_workbook()
            sheet = ExcelMovementService._get_or_create_sheet(workbook)

            ExcelMovementService._ensure_header(sheet)

            loan_items = ExcelMovementService._get_loan_items(loan)

            rows_written = 0

            if not loan_items:
                row = ExcelMovementService._build_row(
                    loan=loan,
                    item=None,
                    movement_type=movement_type,
                    performed_by=performed_by,
                    notes=notes,
                )

                sheet.append(row)
                rows_written += 1

            else:
                for item in loan_items:
                    row = ExcelMovementService._build_row(
                        loan=loan,
                        item=item,
                        movement_type=movement_type,
                        performed_by=performed_by,
                        notes=notes,
                    )

                    sheet.append(row)
                    rows_written += 1

            workbook.save(file_path)

            return {
                "success": True,
                "rows": rows_written,
                "file": file_path,
            }

        finally:
            if workbook is not None:
                workbook.close()

    @staticmethod
    def try_append_loan_movement(
        loan,
        movement_type: str,
        performed_by: str = "",
        notes: str = "",
    ) -> dict:
        """
        Versão segura.

        Se falhar ao escrever na planilha, registra auditoria e não quebra
        a operação principal do sistema.
        """

        try:
            return ExcelMovementService.append_loan_movement(
                loan=loan,
                movement_type=movement_type,
                performed_by=performed_by,
                notes=notes,
            )

        except Exception as exc:
            AuditService.register(
                entity_type="LOAN",
                entity_id=getattr(loan, "id", 0) or 0,
                action="POWERBI_MOVEMENT_EXPORT_FAILED",
                performed_by=performed_by or "SISTEMA",
                new_data={
                    "movement_type": movement_type,
                    "error": str(exc),
                },
            )

            return {
                "success": False,
                "error": str(exc),
            }
    
    @staticmethod
    def _get_movements_file_path() -> str:
        """
        Retorna o caminho da planilha de movimentações usada pelo Power BI.

        Prioridade:
        1. Valor salvo na tela Configurações;
        2. Valor padrão do config.py.
        """

        file_path = AppSettingService.get(
            "EXCEL_MOVEMENTS_FILE",
            current_app.config["EXCEL_MOVEMENTS_FILE"],
        )

        return str(file_path or "").strip()

    @staticmethod
    def _get_movements_sheet_name() -> str:
        """
        Retorna o nome da aba da planilha de movimentações.

        Prioridade:
        1. Valor salvo na tela Configurações;
        2. Valor padrão do config.py.
        """

        sheet_name = AppSettingService.get(
            "EXCEL_MOVEMENTS_SHEET",
            current_app.config["EXCEL_MOVEMENTS_SHEET"],
        )

        return str(sheet_name or "").strip()