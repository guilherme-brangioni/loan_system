import os
import shutil
import tempfile

from datetime import date, datetime
from typing import Any
from flask import current_app
from openpyxl import load_workbook
from database.database import db

from enums.equipment_status import EquipmentStatus

from models.equipment import Equipment

from services.audit_service import AuditService
from services.app_setting_service import AppSettingService

from utils.normalize import normalize_name


class ExcelInventoryService:
    """
    Serviço de sincronização entre a planilha Excel em rede e o banco local.

    Regra principal:
    - A planilha alimenta o banco.
    - O banco controla os empréstimos.
    - Ao cadastrar item manualmente, salvamos primeiro no banco e depois
      tentamos atualizar a planilha.
    """

    REQUIRED_COLUMNS = [
        "REGIONAL",
        "DATA DO INVENTÁRIO",
        "DESCRIÇÃO DO EQUIPAMENTO",
        "MODELO DO EQUIPAMENTO",
        "FABRICANTE",
        "CÓDIGO DO EQUIPAMENTO",
        "NÚMERO DE SÉRIE",
        "STATUS",
        "LOCAL DE ARMAZENAGEM",
        "SUBESTAÇÃO DE ORIGEM",
        "OBSERVAÇÃO",
        "EMPRESTADO PARA",
        "DATA DE EMPRÉSTIMO",
    ]

    # ---------------------------------------------------------------------
    # Funções básicas de arquivo e planilha
    # ---------------------------------------------------------------------

    @staticmethod
    def _copy_network_file_to_temp(source_path: str) -> str:
        """
        Copia a planilha da rede para um arquivo temporário local.

        Usado na leitura/sincronização para evitar manter o arquivo de rede
        aberto durante todo o processamento.
        """

        if not os.path.exists(source_path):
            raise FileNotFoundError(
                f"Planilha não encontrada: {source_path}"
            )

        temp_dir = tempfile.gettempdir()

        temp_file = os.path.join(
            temp_dir,
            f"inventario_sync_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

        shutil.copy2(source_path, temp_file)

        return temp_file

    @staticmethod
    def _load_workbook_from_network_copy():
        """
        Abre uma cópia local da planilha em rede para leitura.
        """

        source_path = AppSettingService.get(
            "EXCEL_INVENTORY_FILE",
            current_app.config["EXCEL_INVENTORY_FILE"],
        )

        temp_file = ExcelInventoryService._copy_network_file_to_temp(
            source_path
        )

        workbook = load_workbook(
            temp_file,
            data_only=True,
            read_only=False,
        )

        return workbook, temp_file

    @staticmethod
    def _load_workbook_for_write():
        """
        Abre a planilha original em rede para escrita.

        Usado quando um item é cadastrado manualmente no sistema e queremos
        refletir essa inclusão/alteração na planilha.

        Atenção:
        se a planilha estiver aberta ou bloqueada, pode ocorrer erro.
        """

        file_path = current_app.config["EXCEL_INVENTORY_FILE"]

        if not os.path.exists(file_path):
            raise FileNotFoundError(
                f"Planilha não encontrada: {file_path}"
            )

        workbook = load_workbook(
            file_path,
            data_only=False,
            read_only=False,
        )

        return workbook, file_path

    # ---------------------------------------------------------------------
    # Funções auxiliares de cabeçalho, célula e data
    # ---------------------------------------------------------------------

    @staticmethod
    def _normalize_header(value: Any) -> str:
        """
        Normaliza o nome das colunas da planilha.
        """

        return normalize_name(str(value or ""))

    @staticmethod
    def _get_headers(sheet) -> dict[str, int]:
        """
        Lê os cabeçalhos da primeira linha da planilha.

        Retorna:
        {
            "REGIONAL": 1,
            "DATA DO INVENTÁRIO": 2,
            ...
        }
        """

        headers: dict[str, int] = {}

        for col_index, cell in enumerate(sheet[1], start=1):
            header = ExcelInventoryService._normalize_header(cell.value)

            if header:
                headers[header] = col_index

        return headers

    @staticmethod
    def _validate_headers(headers: dict[str, int]) -> None:
        """
        Verifica se a planilha possui todas as colunas esperadas.
        """

        missing = [
            column for column in ExcelInventoryService.REQUIRED_COLUMNS
            if column not in headers
        ]

        if missing:
            raise ValueError(
                "A planilha está sem as colunas obrigatórias: "
                + ", ".join(missing)
            )

    @staticmethod
    def _get_cell_value(row, headers: dict[str, int], column_name: str) -> str:
        """
        Lê o valor textual de uma célula pelo nome da coluna.
        """

        col_index = headers.get(column_name)

        if not col_index:
            return ""

        value = row[col_index - 1].value

        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def _get_cell_raw_value(row, headers: dict[str, int], column_name: str):
        """
        Lê o valor bruto da célula.

        Útil para datas, porque o Excel pode devolver datetime/date.
        """

        col_index = headers.get(column_name)

        if not col_index:
            return None

        return row[col_index - 1].value

    @staticmethod
    def _set_cell_if_column_exists(
        sheet,
        headers: dict[str, int],
        row_index: int,
        column_name: str,
        value,
    ) -> None:
        """
        Escreve em uma célula se a coluna existir.

        Isso evita quebra caso alguma coluna opcional mude futuramente.
        """

        col_index = headers.get(column_name)

        if not col_index:
            return

        sheet.cell(row=row_index, column=col_index).value = value

    @staticmethod
    def _parse_excel_date(value) -> datetime | None:
        """
        Converte data vinda do Excel.

        Pode vir como:
        - datetime;
        - date;
        - string;
        - vazio.
        """

        if value is None or value == "":
            return None

        if isinstance(value, datetime):
            return value

        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())

        text = str(value).strip()

        if not text:
            return None

        formats = [
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass

        return None

    # ---------------------------------------------------------------------
    # Funções de identificação e montagem de nome
    # ---------------------------------------------------------------------

    @staticmethod
    def _build_equipment_name(
        fabricante: str,
        modelo: str,
        tipo_equipamento: str,
    ) -> str:
        """
        Monta o nome exibido do equipamento.
        """

        fabricante = normalize_name(fabricante)
        modelo = normalize_name(modelo)
        tipo_equipamento = normalize_name(tipo_equipamento)

        equipment_name = " ".join(
            part for part in [fabricante, modelo]
            if part
        )

        if tipo_equipamento and equipment_name:
            return f"{tipo_equipamento} - {equipment_name}"

        if tipo_equipamento:
            return tipo_equipamento

        return equipment_name or "EQUIPAMENTO SEM DESCRIÇÃO"

    @staticmethod
    def _find_existing_equipment(
        serial: str,
        codigo_equipamento: str,
    ) -> Equipment | None:
        """
        Busca equipamento existente no banco.

        Como a planilha não possui patrimônio, usamos:
        1. Número de Série;
        2. Código do Equipamento.
        """

        serial = normalize_name(serial)
        codigo_equipamento = normalize_name(codigo_equipamento)

        if serial:
            existing = Equipment.query.filter_by(
                serial=serial
            ).first()

            if existing:
                return existing

        if codigo_equipamento:
            existing = Equipment.query.filter_by(
                codigo_equipamento=codigo_equipamento
            ).first()

            if existing:
                return existing

        return None

    @staticmethod
    def _find_row_by_equipment(sheet, headers: dict[str, int], equipment) -> int | None:
        """
        Procura a linha do equipamento na planilha.

        Ordem:
        1. Número de Série;
        2. Código do Equipamento.
        """

        serial_target = normalize_name(getattr(equipment, "serial", "") or "")
        codigo_target = normalize_name(
            getattr(equipment, "codigo_equipamento", "") or ""
        )

        serial_col = headers.get("NÚMERO DE SÉRIE")
        codigo_col = headers.get("CÓDIGO DO EQUIPAMENTO")

        for row_index in range(2, sheet.max_row + 1):
            serial_value = ""
            codigo_value = ""

            if serial_col:
                serial_value = normalize_name(
                    str(sheet.cell(row=row_index, column=serial_col).value or "")
                )

            if codigo_col:
                codigo_value = normalize_name(
                    str(sheet.cell(row=row_index, column=codigo_col).value or "")
                )

            if serial_target and serial_value == serial_target:
                return row_index

            if codigo_target and codigo_value == codigo_target:
                return row_index

        return None

    # ---------------------------------------------------------------------
    # Sincronização: Excel -> Banco
    # ---------------------------------------------------------------------

    @staticmethod
    def sync_inventory_from_excel() -> dict:
        """
        Sincroniza a planilha para o banco local.

        Não apaga equipamentos do banco.
        Apenas cria ou atualiza.
        """

        workbook = None
        temp_file = None

        created = 0
        updated = 0
        ignored = 0
        errors: list[str] = []

        try:
            workbook, temp_file = (
                ExcelInventoryService._load_workbook_from_network_copy()
            )

            sheet_name = AppSettingService.get(
                "EXCEL_INVENTORY_SHEET",
                current_app.config["EXCEL_INVENTORY_SHEET"],
            )

            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Aba '{sheet_name}' não encontrada na planilha."
                )

            sheet = workbook[sheet_name]

            headers = ExcelInventoryService._get_headers(sheet)

            ExcelInventoryService._validate_headers(headers)

            for row_index, row in enumerate(
                sheet.iter_rows(min_row=2),
                start=2,
            ):
                regional = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "REGIONAL",
                )

                data_inventario_raw = ExcelInventoryService._get_cell_raw_value(
                    row,
                    headers,
                    "DATA DO INVENTÁRIO",
                )

                tipo_equipamento = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "DESCRIÇÃO DO EQUIPAMENTO",
                )

                modelo = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "MODELO DO EQUIPAMENTO",
                )

                fabricante = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "FABRICANTE",
                )

                codigo_equipamento = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "CÓDIGO DO EQUIPAMENTO",
                )

                serial = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "NÚMERO DE SÉRIE",
                )

                status_planilha = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "STATUS",
                )

                local_armazenagem = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "LOCAL DE ARMAZENAGEM",
                )

                subestacao_origem = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "SUBESTAÇÃO DE ORIGEM",
                )

                observacao = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "OBSERVAÇÃO",
                )

                emprestado_para = ExcelInventoryService._get_cell_value(
                    row,
                    headers,
                    "EMPRESTADO PARA",
                )

                data_emprestimo_raw = ExcelInventoryService._get_cell_raw_value(
                    row,
                    headers,
                    "DATA DE EMPRÉSTIMO",
                )

                regional = normalize_name(regional)
                tipo_equipamento = normalize_name(tipo_equipamento)
                modelo = normalize_name(modelo)
                fabricante = normalize_name(fabricante)
                codigo_equipamento = normalize_name(codigo_equipamento)
                serial = normalize_name(serial)
                status_planilha = normalize_name(status_planilha)
                local_armazenagem = normalize_name(local_armazenagem)
                subestacao_origem = normalize_name(subestacao_origem)

                if not any(
                    [
                        regional,
                        tipo_equipamento,
                        modelo,
                        fabricante,
                        codigo_equipamento,
                        serial,
                    ]
                ):
                    ignored += 1
                    continue

                if not serial and not codigo_equipamento:
                    errors.append(
                        f"Linha {row_index}: sem Número de Série e sem Código do Equipamento."
                    )
                    continue

                if not tipo_equipamento:
                    errors.append(
                        f"Linha {row_index}: Descrição do Equipamento vazia."
                    )
                    continue

                nome = ExcelInventoryService._build_equipment_name(
                    fabricante=fabricante,
                    modelo=modelo,
                    tipo_equipamento=tipo_equipamento,
                )

                equipment = ExcelInventoryService._find_existing_equipment(
                    serial=serial,
                    codigo_equipamento=codigo_equipamento,
                )

                data_inventario = ExcelInventoryService._parse_excel_date(
                    data_inventario_raw
                )

                data_emprestimo_planilha = (
                    ExcelInventoryService._parse_excel_date(
                        data_emprestimo_raw
                    )
                )

                if equipment:
                    equipment.nome = nome
                    equipment.nome_normalizado = normalize_name(nome)

                    equipment.fabricante = fabricante
                    equipment.modelo = modelo
                    equipment.tipo_equipamento = tipo_equipamento

                    equipment.codigo_equipamento = codigo_equipamento or None
                    equipment.serial = serial or None

                    # A planilha não tem patrimônio.
                    # Não sobrescrevemos patrimônio existente se já houver.
                    if not equipment.patrimonio:
                        equipment.patrimonio = "N/A"

                    equipment.regional = regional
                    equipment.data_inventario = data_inventario
                    equipment.status_planilha = status_planilha
                    equipment.local_armazenagem = local_armazenagem
                    equipment.subestacao_origem = subestacao_origem
                    equipment.emprestado_para_planilha = emprestado_para
                    equipment.data_emprestimo_planilha = data_emprestimo_planilha

                    equipment.observacoes = observacao
                    equipment.validado = True

                    updated += 1

                else:
                    equipment = Equipment()

                    equipment.codigo_interno = "TEMP"

                    equipment.nome = nome
                    equipment.nome_normalizado = normalize_name(nome)

                    equipment.fabricante = fabricante
                    equipment.modelo = modelo
                    equipment.tipo_equipamento = tipo_equipamento

                    # A planilha não possui patrimônio.
                    equipment.patrimonio = "N/A"

                    equipment.codigo_equipamento = codigo_equipamento or None
                    equipment.serial = serial or None

                    equipment.regional = regional
                    equipment.data_inventario = data_inventario
                    equipment.status_planilha = status_planilha
                    equipment.local_armazenagem = local_armazenagem
                    equipment.subestacao_origem = subestacao_origem
                    equipment.emprestado_para_planilha = emprestado_para
                    equipment.data_emprestimo_planilha = data_emprestimo_planilha

                    equipment.observacoes = observacao
                    equipment.validado = True
                    equipment.status = EquipmentStatus.DISPONIVEL.value

                    db.session.add(equipment)
                    db.session.flush()

                    equipment.codigo_interno = f"EQP-{equipment.id:04d}"

                    created += 1

            db.session.commit()

            AuditService.register(
                entity_type="EXCEL_SYNC",
                entity_id=0,
                action="SYNC_INVENTORY_FROM_EXCEL",
                performed_by="SISTEMA",
                new_data={
                    "created": created,
                    "updated": updated,
                    "ignored": ignored,
                    "errors": errors,
                },
            )

            return {
                "created": created,
                "updated": updated,
                "ignored": ignored,
                "errors": errors,
            }

        except Exception:
            db.session.rollback()
            raise

        finally:
            if workbook is not None:
                workbook.close()

            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except OSError:
                    pass

    # ---------------------------------------------------------------------
    # Atualização: Banco -> Excel
    # ---------------------------------------------------------------------

    @staticmethod
    def upsert_equipment_to_excel(equipment) -> dict:
        """
        Atualiza ou adiciona um equipamento na planilha de inventário.

        Usado quando o item é cadastrado manualmente no sistema.

        Importante:
        - O banco já deve ter sido atualizado antes.
        - Se a planilha falhar, o cadastro no banco permanece.
        - Patrimônio não é gravado porque a planilha atual não possui coluna
          de patrimônio.
        """

        workbook = None

        try:
            workbook, file_path = ExcelInventoryService._load_workbook_for_write()

            sheet_name = AppSettingService.get(
                "EXCEL_INVENTORY_SHEET",
                current_app.config["EXCEL_INVENTORY_SHEET"],
            )

            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Aba '{sheet_name}' não encontrada na planilha."
                )

            sheet = workbook[sheet_name]

            headers = ExcelInventoryService._get_headers(sheet)

            ExcelInventoryService._validate_headers(headers)

            row_index = ExcelInventoryService._find_row_by_equipment(
                sheet=sheet,
                headers=headers,
                equipment=equipment,
            )

            action = "updated"

            if row_index is None:
                row_index = sheet.max_row + 1
                action = "created"

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "REGIONAL",
                getattr(equipment, "regional", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "DATA DO INVENTÁRIO",
                getattr(equipment, "data_inventario", None) or datetime.now(),
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "DESCRIÇÃO DO EQUIPAMENTO",
                getattr(equipment, "tipo_equipamento", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "MODELO DO EQUIPAMENTO",
                getattr(equipment, "modelo", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "FABRICANTE",
                getattr(equipment, "fabricante", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "CÓDIGO DO EQUIPAMENTO",
                getattr(equipment, "codigo_equipamento", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "NÚMERO DE SÉRIE",
                getattr(equipment, "serial", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "STATUS",
                getattr(equipment, "status", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "LOCAL DE ARMAZENAGEM",
                getattr(equipment, "local_armazenagem", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "SUBESTAÇÃO DE ORIGEM",
                getattr(equipment, "subestacao_origem", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "OBSERVAÇÃO",
                getattr(equipment, "observacoes", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "EMPRESTADO PARA",
                getattr(equipment, "emprestado_para_planilha", "") or "",
            )

            ExcelInventoryService._set_cell_if_column_exists(
                sheet,
                headers,
                row_index,
                "DATA DE EMPRÉSTIMO",
                getattr(equipment, "data_emprestimo_planilha", None) or "",
            )

            workbook.save(file_path)

            AuditService.register(
                entity_type="EQUIPMENT",
                entity_id=getattr(equipment, "id", 0) or 0,
                action="UPSERT_EQUIPMENT_TO_EXCEL",
                performed_by="SISTEMA",
                new_data={
                    "action": action,
                    "codigo_interno": getattr(equipment, "codigo_interno", ""),
                    "codigo_equipamento": getattr(
                        equipment,
                        "codigo_equipamento",
                        "",
                    ),
                    "serial": getattr(equipment, "serial", ""),
                    "excel_row": row_index,
                },
            )

            return {
                "success": True,
                "action": action,
                "row": row_index,
            }

        finally:
            if workbook is not None:
                workbook.close()